import pandas as pd
import openpyxl

def load_excel_files(query_report_file, scs_file):
    df1 = pd.read_excel(query_report_file)
    df2 = pd.read_excel(scs_file)

    return df1, df2

def main():
    query_report_file = "./docs/QueryReport.xlsx"
    scs_file = "./docs/SCS.xlsx"

    df1, df2 = load_excel_files(query_report_file, scs_file)

    df1.reset_index(drop=True, inplace=True)
    df1.to_excel("PCCS.xlsx", index=False)

    with openpyxl.load_workbook("PCCS.xlsx") as wb:
        oli_sheet = wb["oli"]

        sku_data = [cell.value for cell in oli_sheet["A"][1:]]
        combined_data = [str(d.value) + " " + str(e.value) for d, e in zip(oli_sheet["D"][1:], oli_sheet["E"][1:])]
        series_data = [cell.value for cell in oli_sheet["F"][1:]]

        for i, sku in enumerate(sku_data):
            osinstalled = df2.loc[(df2["SKU"] == sku) & (df2["ContainerName"] == "osinstalled")]
            processorname = df2.loc[(df2["SKU"] == sku) & (df2["ContainerName"] == "processorname")]
            memstdes_01 = df2.loc[(df2["SKU"] == sku) & (df2["ContainerName"] == "memstdes_01")]
            hd_01des = df2.loc[(df2["SKU"] == sku) & (df2["ContainerName"] == "hd_01des")]

            oli_sheet.cell(row=4, column=i).value = osinstalled.iloc[0]["ContainerValue"]
            oli_sheet.cell(row=5, column=i).value = processorname.iloc[0]["ContainerValue"]
            oli_sheet.cell(row=6, column=i).value = memstdes_01.iloc[0]["ContainerValue"]
            oli_sheet.cell(row=7, column=i).value = hd_01des.iloc[0]["ContainerValue"]

    wb.save("PCCS.xlsx")

if __name__ == "__main__":
    main()
