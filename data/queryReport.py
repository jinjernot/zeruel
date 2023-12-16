import pandas as pd
from openpyxl import load_workbook


def queryReport(file1):

    try:

        df = pd.read_excel(file1)
        df.to_excel('QueryReport.xlsx', index=False)
        wb = load_workbook(file1)
        ws = wb.active
        ws = wb.create_sheet(title='CAS')  

        sku_values = df['Sku'].values
        oid_values = df['OID'].values
        
        for i in range(len(sku_values)):
            ws.cell(row=i + 1, column=1).value = sku_values[i]
            ws.cell(row=i + 1, column=2).value = oid_values[i]


        wb.save('QueryReport.xlsx')

    except Exception as e:
        print(e)
    return 