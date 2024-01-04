import pandas as pd
from openpyxl import load_workbook

def load_excel_files(file1, file2):
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    return df1, df2

def main():
    file1 = "./docs/QueryReport.xlsx"
    file2 = "./docs/SCS.xlsx"

    df1, df2 = load_excel_files(file1, file2)

    # Load the first Excel file into a DataFrame.
    df1 = pd.read_excel(file1)

    # Remove the index from the DataFrame.
    df1.reset_index(drop=True, inplace=True)

    # Save the DataFrame to a new Excel file.
    df1.to_excel("PCCS.xlsx", index=False)

    wb = load_workbook("PCCS.xlsx")

    # Set "Sheet1" as the active sheet.
    wb.active = wb["Sheet1"]  
    sheet1 = wb.active

    # Get the content from column A.
    sku_column = sheet1["A"]
    sku_data = [cell.value for cell in sku_column]

    # Get the content from columns D and E.
    column_d = sheet1["D"]
    column_e = sheet1["E"]
    combined_data = [str(d.value) + " " + str(e.value) for d, e in zip(column_d, column_e)]

    # Get the content from column F.
    series_column = sheet1["F"]
    series_data = [cell.value for cell in series_column]

    # Create the "oli" sheet and paste the data into new columns.
    wb.create_sheet("oli")
    oli_sheet = wb["oli"]

    for i in range(1, len(sku_data)):
        oli_sheet.cell(row=1, column=i).value = sku_data[i]
    for i in range(1, len(combined_data)):
        oli_sheet.cell(row=2, column=i).value = combined_data[i]
    for i in range(1, len(sku_data)):
        oli_sheet.cell(row=3, column=i).value = series_data[i]


    # Iterate across all values from sku_data starting from the second value.
    for i in range(1, len(sku_data)):
        sku_search_value = sku_data[i]

        # Perform the search in "df2" using the selected value.
        osinstalled = df2[(df2["SKU"] == sku_search_value) & (df2["ContainerName"] == "osinstalled")]
        processorname = df2[(df2["SKU"] == sku_search_value) & (df2["ContainerName"] == "processorname")]
        memstdes_01 = df2[(df2["SKU"] == sku_search_value) & (df2["ContainerName"] == "memstdes_01")]
        hd_01des = df2[(df2["SKU"] == sku_search_value) & (df2["ContainerName"] == "hd_01des")]

        # Retrieve the values from the "ContainerValue" column of the search results.
        osinstalled_values = osinstalled["ContainerValue"].values.tolist()
        processorname_values = processorname["ContainerValue"].values.tolist()
        memstdes_01_values = memstdes_01["ContainerValue"].values.tolist()
        hd_01des_01_values = hd_01des["ContainerValue"].values.tolist()

        for j, value in enumerate(osinstalled_values, start=1):
            oli_sheet.cell(row=4, column=i).value = value
        for j, value in enumerate(processorname_values, start=1):
            oli_sheet.cell(row=5, column=i).value = value
        for j, value in enumerate(memstdes_01_values, start=1):
            oli_sheet.cell(row=6, column=i).value = value
        for j, value in enumerate(hd_01des_01_values, start=1):
            oli_sheet.cell(row=7, column=i).value = value

    # Save the modified Excel file.
    wb.save("PCCS.xlsx")

if __name__ == "__main__":
    main()