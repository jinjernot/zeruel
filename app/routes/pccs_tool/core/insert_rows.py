from app.routes.pccs_tool.core.get_container_value import *
from app.routes.pccs_tool.core.get_product import *
from openpyxl.styles import Font, PatternFill

import pandas as pd
import json
import io

from config import WARRANTY_PATH

def insert_rows(file, output_buffer=None):
    """
    Create a new dataframe with the values found in both reports, insert the productlongname and warranty values.
    """
    # Read the file content into a BytesIO object
    file_content = io.BytesIO(file.read())

    # Read the sheets into dataframes
    df1 = pd.read_excel(file_content, sheet_name='PRISM QUERY', engine='openpyxl')
    file_content.seek(0)  # Reset the BytesIO object to the beginning
    df2 = pd.read_excel(file_content, sheet_name='SCS', engine='openpyxl')
    file_content.seek(0)  # Reset the BytesIO object to the beginning again

    merged_df = pd.merge(df1, df2, left_on="Sku", right_on="SKU", how="inner")

    rows_to_insert = []
    processed_skus = set()
    errors = []

    for _, row in merged_df.iterrows():
        sku = row["Sku"]

        if sku in processed_skus:
            continue

        osinstalled_value = get_container_value(df2, sku, "osinstalled")
        processorname_value = get_container_value(df2, sku, "processorname")
        memstdes_01_value = get_container_value(df2, sku, "memstdes_01")
        hd_01des_value = get_container_value(df2, sku, "hd_01des")
        hd_02des_value = get_container_value(df2, sku, "hd_02des")
        hd_03des_value = get_container_value(df2, sku, "hd_03des")

        reason = []

        if osinstalled_value is None:
            reason.append("osinstalled is None")
        if processorname_value is None:
            reason.append("processorname is None")
        if memstdes_01_value is None:
            reason.append("memstdes_01 is None")
        if hd_01des_value is None:
            reason.append("hd_01des is None")
        if hd_01des_value is not None and hd_01des_value.strip() == "":
            reason.append("hd_01des is blank")

        if reason:
            name_value = row["Name"]
            chunk_value = f"{name_value} - Unprocessed due to: {', '.join(reason)}"
            proddiff_long_row = {
                "Sku": sku,
                "Item_Id": row["Item_Id"],
                "ItemLevel": "Product",
                "CultureCode": "na-en",
                "DataType": "Text",
                "Tag": "proddiff_long",
                "ChunkValue": chunk_value
            }
            rows_to_insert.append(proddiff_long_row)
            errors.append(proddiff_long_row)
        else:
            name_value = row["Name"]
            chunk_value = f"{name_value} with {osinstalled_value}, {processorname_value}, {memstdes_01_value}"

            if hd_01des_value not in ("[BLANK]", None):
                chunk_value += f", {hd_01des_value}"
            if hd_02des_value not in ("[BLANK]", None):
                chunk_value += f", {hd_02des_value}"
            if hd_03des_value not in ("[BLANK]", None):
                chunk_value += f", {hd_03des_value}"

            chunk_value += ", Low halogen"
            
            api_data = get_product(sku)

            if api_data is not None:
                proddiff_long_row = {
                    "Sku": sku,
                    "Item_Id": row["Item_Id"],
                    "ItemLevel": "Product",
                    "CultureCode": "na-en",
                    "DataType": "Text",
                    "Tag": "proddiff_long",
                    "ChunkValue": chunk_value
                }
                rows_to_insert.append(proddiff_long_row)

                if isinstance(api_data, dict):
                    chunk_value = api_data.get("name")
                else:
                    chunk_value = api_data

                warranty_row = {
                    "Sku": sku,
                    "Item_Id": row["Item_Id"],
                    "ItemLevel": "Product",
                    "CultureCode": "na-en",
                    "DataType": "Text",
                    "Tag": "wrntyfeatures",
                    "ChunkValue": chunk_value
                }
                rows_to_insert.append(warranty_row)

                processed_skus.add(sku)

    # Remove duplicate rows
    pccs_df = pd.DataFrame(rows_to_insert).drop_duplicates()

    # Load the MS4 sheet and warranty data
    file_content.seek(0)  # Reset the BytesIO object to the beginning again
    ms4_filtered_df = pd.read_excel(file_content, sheet_name='MS4', engine='openpyxl')
    
    # Clean up SKU values in ms4_filtered_df
    ms4_filtered_df['SKU                                     '] = ms4_filtered_df['SKU                                     '].str.replace(r'#.*$', '', regex=True)

    with open(WARRANTY_PATH, 'r') as json_file:
        warranty_data = json.load(json_file)

    # Iterate through pccs_df and check warranty information
    for index, row in pccs_df.iterrows():
        sku = row["Sku"]
        chunk_value = row["ChunkValue"]
        tag = row["Tag"]

        # Validate if SKU exists in both ms4_filtered_df and merged_df
        if tag == "wrntyfeatures":
            if sku in ms4_filtered_df["SKU                                     "].values and sku in merged_df["Sku"].values:
                # Find matching product in warranty data
                matching_products = [product for warranty_list in warranty_data["warranty"] for product in warranty_list if product["Product"] == chunk_value]
                
                if matching_products:
                    description = matching_products[0]["Description"]
                    
                    warranty_value = matching_products[0]["Warranty"]
                    if any(ms4_filtered_df["DESCRIPTION                             "].str.contains(warranty_value)):
                        pccs_df.at[index, "ChunkValue"] = description
                else:
                    # If no matching product found, append the reason to the existing ChunkValue
                    reason = "No warranty information available"
                    pccs_df.at[index, "ChunkValue"] = f"{chunk_value}, {reason}"
            else:
                # If SKU does not exist in either DataFrame, append the reason
                reason = "SKU not found in MS4 report"
                pccs_df.at[index, "ChunkValue"] = f"{sku}, {reason}"

    pccs_df['ChunkStatus'] = 'F'
    pccs_df['SourceLevel'] = ''
    pccs_df['SourceCulture'] = ''

    if output_buffer:
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            pccs_df.to_excel(writer, index=False, sheet_name='CASChunks')

            workbook = writer.book

            for sheet in workbook.sheetnames:
                worksheet = writer.sheets[sheet]
                for cell in worksheet["1:1"]:
                    cell.font = Font(bold=True)

            # Add errors in red font
            error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            error_font = Font(color="9C0006")

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    if "Unprocessed due to" in str(cell.value) or "No warranty information available" in str(cell.value) or "SKU not found in MS4 report" in str(cell.value):
                        cell.fill = error_fill
                        cell.font = error_font